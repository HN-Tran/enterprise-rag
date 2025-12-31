"""Extractors for PDF/DOCX/XLSX/HTML/ASPX into logical pages."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List

import trafilatura
from bs4 import BeautifulSoup

from app.ingestion.citations import EmbeddedLink


@dataclass(frozen=True)
class ExtractedDoc:
    title: str
    source_type: str
    uri: str
    pages: List[str]
    links: List[EmbeddedLink] = field(default_factory=list)


def extract_pdf(path: str) -> ExtractedDoc:
    import fitz  # PyMuPDF

    p = Path(path)
    doc = fitz.open(path)
    pages = []
    links = []

    for page_no, page in enumerate(doc, start=1):
        pages.append((page.get_text("text") or "").strip())

        # Extract embedded hyperlinks
        for link in page.get_links():
            uri = link.get("uri")
            if uri and uri.startswith(("http://", "https://")):
                # Try to get link text from the rect area
                rect = link.get("from")
                link_text = None
                if rect:
                    try:
                        link_text = page.get_text("text", clip=rect).strip()
                    except Exception:
                        pass
                links.append(EmbeddedLink(
                    page_no=page_no,
                    uri=uri,
                    text=link_text,
                ))

    doc.close()
    return ExtractedDoc(
        title=p.stem,
        source_type="pdf",
        uri=str(p.resolve()),
        pages=pages,
        links=links,
    )


def extract_docx(path: str) -> ExtractedDoc:
    from docx import Document

    p = Path(path)
    d = Document(path)
    sections: List[str] = []
    buf: list[str] = []

    for para in d.paragraphs:
        t = (para.text or "").strip()
        if not t:
            continue
        style = (para.style.name or "").lower() if para.style else ""
        is_heading = "heading" in style
        if is_heading and buf:
            sections.append("\n".join(buf).strip())
            buf = [t]
        else:
            buf.append(t)

    if buf:
        sections.append("\n".join(buf).strip())

    if not sections:
        sections = [""]

    return ExtractedDoc(title=p.stem, source_type="docx", uri=str(p.resolve()), pages=sections)


def extract_xlsx(path: str) -> ExtractedDoc:
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(path, data_only=True)
    pages: List[str] = []

    for sheet in wb.worksheets:
        rows_out = []
        max_rows = min(sheet.max_row, 2000)
        max_cols = min(sheet.max_column, 80)

        for r in range(1, max_rows + 1):
            vals = []
            empty = True
            for c in range(1, max_cols + 1):
                v = sheet.cell(r, c).value
                s = "" if v is None else str(v).strip()
                if s:
                    empty = False
                vals.append(s)
            if empty:
                continue
            rows_out.append("\t".join(vals))
            if len(rows_out) >= 400:
                break

        text = f"Sheet: {sheet.title}\n" + ("\n".join(rows_out) if rows_out else "")
        pages.append(text.strip())

    if not pages:
        pages = [""]

    return ExtractedDoc(title=p.stem, source_type="xlsx", uri=str(p.resolve()), pages=pages)


def extract_html(path: str) -> ExtractedDoc:
    p = Path(path)
    html = p.read_text(encoding="utf-8", errors="ignore")

    extracted = trafilatura.extract(html, include_tables=True) or ""
    if not extracted.strip():
        soup = BeautifulSoup(html, "lxml")
        extracted = soup.get_text("\n")

    # Heuristic segmentation into logical pages
    parts = [x.strip() for x in extracted.split("\n\n") if x.strip()]
    pages: List[str] = []
    buf: list[str] = []
    size = 0

    for part in parts:
        if size + len(part) > 12000 and buf:
            pages.append("\n\n".join(buf).strip())
            buf, size = [], 0
        buf.append(part)
        size += len(part)

    if buf:
        pages.append("\n\n".join(buf).strip())

    if not pages:
        pages = [""]

    return ExtractedDoc(title=p.stem, source_type="html", uri=str(p.resolve()), pages=pages)


def extract_any(path: str) -> ExtractedDoc:
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".docx":
        return extract_docx(path)
    if ext in (".xlsx", ".xlsm"):
        return extract_xlsx(path)
    if ext in (".html", ".htm", ".aspx"):
        return extract_html(path)
    raise ValueError(f"Unsupported file type: {ext}")
