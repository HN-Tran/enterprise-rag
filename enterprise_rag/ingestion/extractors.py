"""Extractors for PDF/DOCX/XLSX/HTML/ASPX into logical pages."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List

import trafilatura
from bs4 import BeautifulSoup

from enterprise_rag.ingestion.citations import EmbeddedLink


@dataclass(frozen=True)
class ExtractedDoc:
    title: str
    source_type: str
    uri: str
    pages: List[str]
    links: List[EmbeddedLink] = field(default_factory=list)


def extract_xls(path: str) -> ExtractedDoc:
    """
    Enterprise-grade .xls extractor with the same logic as extract_xlsx():
    - header detection (row 1)
    - row formatting: "Header: Value | ..."
    - skips empty rows
    - row/col caps
    - overlapping sliding-window chunks
    - chunk header includes sheet + Zeilen range
    """
    import xlrd

    p = Path(path)
    book = xlrd.open_workbook(path, on_demand=True)
    pages: List[str] = []

    # Keep identical chunk settings to extract_xlsx
    # Each row is an independent record with header context included
    ROWS_PER_CHUNK = 1
    OVERLAP_ROWS = 0

    # Keep identical caps to extract_xlsx
    MAX_ROWS = 5000
    MAX_COLS = 50

    def _cell_to_str(v) -> str:
        # xlrd returns floats for numbers; we keep behavior close to openpyxl's str(v).strip()
        # but avoid "1.0" when the cell is an integer.
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        if isinstance(v, (int,)):
            return str(v)
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            return str(v)
        return str(v).strip()

    try:
        for sheet in book.sheets():
            max_rows = min(sheet.nrows or 0, MAX_ROWS)
            max_cols = min(sheet.ncols or 0, MAX_COLS)

            if max_rows == 0 or max_cols == 0:
                continue

            # Header row assumed to be first row (index 0 in xlrd)
            headers: List[str] = []
            for c in range(max_cols):
                headers.append(_cell_to_str(sheet.cell_value(0, c)))
            has_header = any(headers)
            data_start = 1 if has_header else 0  # xlrd is 0-based

            all_rows: list[tuple[int, str]] = []

            # Iterate data rows, build row_text exactly like extract_xlsx (but with 1-based row numbers)
            for r in range(data_start, max_rows):
                vals: List[str] = []
                empty = True

                for c in range(max_cols):
                    s = _cell_to_str(sheet.cell_value(r, c))
                    if s:
                        empty = False
                    vals.append(s)

                if empty:
                    continue

                if has_header:
                    parts: List[str] = []
                    for c, (h, val) in enumerate(zip(headers, vals), start=1):
                        if not val:
                            continue
                        # Same fallback label semantics as your xlsx extractor
                        label = h if h else f"Spalte{c}"
                        parts.append(f"{label}: {val}")
                    row_text = " | ".join(parts)
                else:
                    row_text = " | ".join(v for v in vals if v)

                if row_text:
                    # Store 1-based Excel row number to match extract_xlsx output
                    all_rows.append((r + 1, row_text))

            if not all_rows:
                continue

            # Sliding-window overlapping chunks
            stride = ROWS_PER_CHUNK - OVERLAP_ROWS
            i = 0
            while i < len(all_rows):
                chunk = all_rows[i : i + ROWS_PER_CHUNK]
                if not chunk:
                    break

                first_row = chunk[0][0]
                last_row = chunk[-1][0]
                chunk_text = f"Sheet: {sheet.name} (Zeilen {first_row}-{last_row})\n"
                chunk_text += "\n".join(row[1] for row in chunk)
                pages.append(chunk_text.strip())

                i += stride
                # Avoid tiny final chunks (same rule as xlsx)
                if i < len(all_rows) and len(all_rows) - i < OVERLAP_ROWS:
                    break
    finally:
        # Ensure file handles are released
        try:
            book.release_resources()
        except Exception:
            pass

    if not pages:
        pages = [""]

    return ExtractedDoc(
        title=p.stem,
        source_type="xls",
        uri=str(p.resolve()),
        pages=pages,
    )


def extract_pdf(path: str) -> ExtractedDoc:
    import fitz  # PyMuPDF

    p = Path(path)
    doc = fitz.open(path)
    pages = []
    links = []

    for page_no, page in enumerate(doc, start=1):
        pages.append((page.get_text("text") or "").strip())

        # Extract embedded hyperlinks
        for link in page.get_links():
            uri = link.get("uri")
            if uri and uri.startswith(("http://", "https://")):
                # Try to get link text from the rect area
                rect = link.get("from")
                link_text = None
                if rect:
                    try:
                        link_text = page.get_text("text", clip=rect).strip()
                    except Exception:
                        pass
                links.append(EmbeddedLink(
                    page_no=page_no,
                    uri=uri,
                    text=link_text,
                ))

    doc.close()
    return ExtractedDoc(
        title=p.stem,
        source_type="pdf",
        uri=str(p.resolve()),
        pages=pages,
        links=links,
    )


def extract_docx(path: str) -> ExtractedDoc:
    from docx import Document

    p = Path(path)
    d = Document(path)
    sections: List[str] = []
    buf: list[str] = []

    for para in d.paragraphs:
        t = (para.text or "").strip()
        if not t:
            continue
        style = (para.style.name or "").lower() if para.style else ""
        is_heading = "heading" in style
        if is_heading and buf:
            sections.append("\n".join(buf).strip())
            buf = [t]
        else:
            buf.append(t)

    if buf:
        sections.append("\n".join(buf).strip())

    if not sections:
        sections = [""]

    return ExtractedDoc(title=p.stem, source_type="docx", uri=str(p.resolve()), pages=sections)


def extract_doc(path: str) -> ExtractedDoc:
    """Extract text from legacy .doc files using antiword or catdoc."""
    import subprocess

    p = Path(path)
    text = ""

    # Try antiword first (better formatting)
    try:
        result = subprocess.run(
            ["antiword", "-w", "0", path],  # -w 0 = no line wrapping
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0:
            text = result.stdout
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    # Fallback to catdoc
    if not text:
        try:
            result = subprocess.run(
                ["catdoc", "-w", path],  # -w = no line wrapping
                capture_output=True,
                text=True,
                timeout=60,
            )
            if result.returncode == 0:
                text = result.stdout
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    if not text:
        raise ValueError(
            f"Cannot extract .doc file: {path}. "
            "Install 'antiword' or 'catdoc': sudo apt install antiword catdoc"
        )

    # Split into logical sections (similar to docx)
    paragraphs = [x.strip() for x in text.split("\n\n") if x.strip()]
    sections: List[str] = []
    buf: list[str] = []
    size = 0

    for para in paragraphs:
        if size + len(para) > 12000 and buf:
            sections.append("\n\n".join(buf).strip())
            buf, size = [], 0
        buf.append(para)
        size += len(para)

    if buf:
        sections.append("\n\n".join(buf).strip())

    if not sections:
        sections = [""]

    return ExtractedDoc(title=p.stem, source_type="doc", uri=str(p.resolve()), pages=sections)


def extract_xlsx(path: str) -> ExtractedDoc:
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(path, data_only=True)
    pages: List[str] = []

    # Each row is an independent record with header context included
    # 1 row per chunk maximizes BM25 term density
    ROWS_PER_CHUNK = 1
    OVERLAP_ROWS = 0

    for sheet in wb.worksheets:
        max_rows = min(sheet.max_row or 0, 5000)
        max_cols = min(sheet.max_column or 0, 50)

        if max_rows == 0 or max_cols == 0:
            continue

        # Extract header row (assumed to be row 1)
        headers = []
        for c in range(1, max_cols + 1):
            v = sheet.cell(1, c).value
            headers.append("" if v is None else str(v).strip())
        has_header = any(headers)
        data_start = 2 if has_header else 1

        # Extract all data rows with header context
        all_rows = []
        for r in range(data_start, max_rows + 1):
            vals = []
            empty = True
            for c in range(1, max_cols + 1):
                v = sheet.cell(r, c).value
                s = "" if v is None else str(v).strip()
                if s:
                    empty = False
                vals.append(s)
            if empty:
                continue

            # Format as "Header1: Value1 | Header2: Value2 | ..."
            if has_header:
                parts = []
                for h, val in zip(headers, vals):
                    if val:  # Only include non-empty values
                        label = h if h else f"Spalte{headers.index(h)+1}"
                        parts.append(f"{label}: {val}")
                row_text = " | ".join(parts)
            else:
                row_text = " | ".join(v for v in vals if v)

            if row_text:
                all_rows.append((r, row_text))

        # Create overlapping chunks
        if not all_rows:
            continue

        stride = ROWS_PER_CHUNK - OVERLAP_ROWS
        i = 0
        while i < len(all_rows):
            chunk = all_rows[i:i + ROWS_PER_CHUNK]
            if not chunk:
                break

            first_row = chunk[0][0]
            last_row = chunk[-1][0]
            chunk_text = f"Sheet: {sheet.title} (Zeilen {first_row}-{last_row})\n"
            chunk_text += "\n".join(row[1] for row in chunk)
            pages.append(chunk_text.strip())

            i += stride
            # Avoid tiny final chunks
            if i < len(all_rows) and len(all_rows) - i < OVERLAP_ROWS:
                break

    if not pages:
        pages = [""]

    return ExtractedDoc(title=p.stem, source_type="xlsx", uri=str(p.resolve()), pages=pages)


def extract_html(path: str) -> ExtractedDoc:
    p = Path(path)
    html = p.read_text(encoding="utf-8", errors="ignore")

    extracted = trafilatura.extract(html, include_tables=True) or ""
    if not extracted.strip():
        soup = BeautifulSoup(html, "lxml")
        extracted = soup.get_text("\n")

    # Heuristic segmentation into logical pages
    parts = [x.strip() for x in extracted.split("\n\n") if x.strip()]
    pages: List[str] = []
    buf: list[str] = []
    size = 0

    for part in parts:
        if size + len(part) > 12000 and buf:
            pages.append("\n\n".join(buf).strip())
            buf, size = [], 0
        buf.append(part)
        size += len(part)

    if buf:
        pages.append("\n\n".join(buf).strip())

    if not pages:
        pages = [""]

    return ExtractedDoc(title=p.stem, source_type="html", uri=str(p.resolve()), pages=pages)


def extract_any(path: str) -> ExtractedDoc:
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".doc":
        return extract_doc(path)
    if ext == ".xls":
        return extract_xls(path)
    if ext in (".xlsx", ".xlsm"):
        return extract_xlsx(path)
    if ext in (".html", ".htm", ".asp", ".aspx"):
        return extract_html(path)
    raise ValueError(f"Unsupported file type: {ext}")
